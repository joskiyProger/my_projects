import tkinter as tk
import threading
import pandas as pd
from tkinter import ttk
from tkinter import *
from datetime import date, datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import os
import re
import shutil
df= None
number_label = None
headers_list = None
table = None
entries = None
root = None
age_label = None
last_name_label_add = None
name_label_add = None
frame_add = None
frame_find = None
cemetery_label = None
tabs_list = None
entries_old = None
frame_add_old = None
number_label_old = None
age_label_old = None
last_name_label_add_old = None
name_label_add_old = None
cemetery_label_old = None

def read_sheet(df):
    wb = load_workbook('load.xlsx')
    ws = wb.active
    first_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):   
        if str(row[0].value) == '1':
            first_row = row[0].row
            break
    empty_rows = []
    for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            empty_rows.append(row)
    for row in empty_rows:
        ws.delete_rows(row[0].row)
    wb.save('load.xlsx')

    df = pd.read_excel('load.xlsx', skiprows=4, engine='openpyxl')
    
    df = {key.capitalize(): value for key, value in df.items()}

    for i in range(4, len(list(df['Имя умершего']))):
        df['Имя умершего'][i] = df['Имя умершего'][i].capitalize()
        df['Фамилия умершего'][i] = df['Фамилия умершего'][i].capitalize()
        if str(df['Отчество умершего'][i]) != 'nan':
            df['Отчество умершего'][i] = df['Отчество умершего'][i].capitalize()

    df = pd.DataFrame(df)
    df = df.dropna(thresh=3)
    df.reset_index(drop=True, inplace=True)  # Сброс текущих индексов
    df.index = df.index + 1  # Установка индексов от 1 до длины датафрейма
    df.set_index('Номер внесения записи')
    tk.Label(frame_find,
            text="Файлы загружаются",
            bg= '#' + str(hex(0xdd6666))[2:],
            font= 20
        ).grid(row=8, column=0, sticky='we', columnspan=2)
    return df

def set_df():
    global df
    global headers_list
    global number_label
    global number_label_old
    global table
    global root
    global frame_add
    global frame_add_old
    global frame_find
    global tabs_list
    # df = pd.DataFrame(df)
    PATH = os.getcwd()
    if not os.path.isdir(PATH + '/copy/'):
        os.mkdir(PATH + '/copy/')
    pattern = '20\d\d_'
    tabs_list = []
    all_files = os.listdir(PATH)

    for file in all_files:
        if re.findall(pattern + 'Покровское.xlsx', file) or re.findall(pattern + 'Южное.xlsx', file) or re.findall(pattern + 'Покровское_\d.xlsx', file) or re.findall(pattern + 'Южное_\d.xlsx', file) or re.findall(pattern + 'Старопокровское.xlsx', file) or re.findall(pattern + 'Старопокровское_\d.xlsx', file):
            tabs_list.append(file)
            this_date = str(datetime.now().year) + '_' + str(str(datetime.now().strftime("%m"))) + '_' + str(datetime.now().strftime("%d")) + '/'
            if not os.path.isdir(PATH + '/copy/' + this_date):
                os.mkdir(PATH + '/copy/' + this_date)
            if os.path.exists(PATH + '/copy/' + this_date + file):
                os.remove(PATH + '/copy/' + this_date + file)
            shutil.copy(PATH + '/' + file, PATH + '/copy/' + this_date + file)
    print(tabs_list)
    df_temp = None
    for tab_name in tabs_list:
        if os.path.exists(PATH + '/load.xlsx'):
            os.remove(PATH + '/load.xlsx')
        shutil.copy(PATH + '/' + tab_name, PATH + '/load.xlsx')
        if df is None:
            df = read_sheet(df)
        else:
            df_temp = read_sheet(df_temp)
            df = pd.concat([df, df_temp], ignore_index=True)    
    headers_list = df.columns.tolist()
    number_label = tk.Label(frame_add,
                 text="Вы сейчас заполняете человека под номером "+str(len(df['Номер внесения записи'])),
                 bg= '#' + str(hex(0xffffff))[2:],
                 font= 20
                )
    number_label.grid(row=8, column=0, sticky='we', columnspan=7)
    number_label_old = tk.Label(frame_add_old,
                 text="Вы сейчас заполняете человека под номером "+str(len(df['Номер внесения записи'])),
                 bg= '#' + str(hex(0xffffff))[2:],
                 font= 20
                )
    number_label_old.grid(row=8, column=0, sticky='we', columnspan=7)
    tk.Label(frame_find,
                 text="Файлы загружены",
                 bg= '#' + str(hex(0xffffff))[2:],
                 font= 20
                ).grid(row=8, column=0, sticky='we', columnspan=2)
    table = ttk.Treeview(columns=headers_list, show="headings")
    xscrollbar = ttk.Scrollbar(root,orient=HORIZONTAL,command=table.xview)
    xscrollbar.grid(row=10, column=0, sticky=EW, columnspan=6)
    table.configure(xscrollcommand=xscrollbar.set)

    yscrollbar = ttk.Scrollbar(root,orient=VERTICAL,command=table.yview)
    yscrollbar.grid(row=9, column=8, sticky=NS)
    table.configure(yscrollcommand=yscrollbar.set)
    for item in headers_list:
        table.column(item, anchor=W, width=len(item) * 11)
        table.heading(item, text=item, anchor=CENTER)
    table.grid(row=9, column=0, sticky="nsew", columnspan=6)

    
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
def change_label_bg_color(parent_func, inserted_data):
    global number_label
    global age_label
    global name_label_add
    global last_name_label_add
    global cemetery_label
    global number_label_old
    global age_label_old
    global name_label_add_old
    global last_name_label_add_old
    global cemetery_label_old
    if parent_func == 'now':
        number_label['text'] = "Вы сейчас заполняете человека под номером "+str(inserted_data[0] + 1)
        age_label['bg'] = '#ffffff'
        name_label_add['bg'] = '#ffffff'
        last_name_label_add['bg'] = '#ffffff'
        cemetery_label['bg'] = '#ffffff'
    if parent_func == 'old':
        number_label_old['text'] = "Вы сейчас заполняете человека под номером "+str(inserted_data[0] + 1)
        age_label_old['bg'] = '#ffffff'
        name_label_add_old['bg'] = '#ffffff'
        last_name_label_add_old['bg'] = '#ffffff'
        cemetery_label_old['bg'] = '#ffffff'

def create_new_excel(parent_func, name, inserted_data):
    global number_label
    global table
    global df
    wb = Workbook()
    ws = wb.active
    for i in range(1, 7):
        ws['A' + str(i)] = 0
    inserted_data[0] = ws['A' + str(ws.max_row)].value + 1
    inserted_data[1] = str(datetime.now().year) + '.' + str(str(datetime.now().strftime("%m"))) + '.' + str(datetime.now().strftime("%d"))
    ws.append(inserted_data)
    col_names = ['Номер внесения записи', 'Дата внесения записи', 'Фамилия умершего', 'Имя умершего', 'Отчество умершего',	'Возраст умершего',	'Дата смерти',	'Номер свидетельства о смерти из ЗАГСа',	'Каким ЗАГСом выдано свидетельство',	'Дата погребения',	'Кладбище',	'Номер сектора могилы', 'Номер ряда могилы', 'Номер места могилы',	'Размер отведенного участка земли для погребения',	'Фамилия землекопа',	'Фамилия, имя, отчество и адрес лица, взявшего на себя обязанность осуществить погребение умершего']
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
    nums = ['1',	'2',	'3',	'4',	'5',	'6',	'7',	'8',	'9',	'10',	'11',	'12',	'13',	'14',	'15',	'16', '17']
    for i in range(0, 17):
        ws[letters[i] + '5'] = col_names[i]
        ws[letters[i] + '6'] = nums[i]
    set_border(ws, 'A5:Q7')
    wb.save(name)
    df.loc[len(df)+1] = inserted_data
    change_label_bg_color(parent_func, inserted_data)
    table.delete(*table.get_children())
    table.insert('', tk.END, values=list(df.iloc[len(df)-1]))

def add_row(parent_func, cemetery, inserted_data, name):
    global age_label
    global name_label_add
    global last_name_label_add
    global number_label
    global number_label_old
    global table
    global tabs_list
    global entries
    
    PATH = os.getcwd()
    pattern = '20\d\d_'
    tabs_list = []
    all_files = os.listdir(PATH)
    for file in all_files:
        if re.findall(pattern + 'Покровское.xlsx', file) or re.findall(pattern + 'Южное.xlsx', file) or re.findall(pattern + 'Покровское_\d.xlsx', file) or re.findall(pattern + 'Южное_\d.xlsx', file) or re.findall(pattern + 'Старопокровское.xlsx', file) or re.findall(pattern + 'Старопокровское_\d.xlsx', file):
            tabs_list.append(file)
    
    if os.path.exists(PATH + '/load.xlsx'):
            os.remove(PATH + '/load.xlsx')
    if not os.path.exists(PATH + '/' + name):
        create_new_excel(parent_func, name, inserted_data)
        return 0
    shutil.copy(PATH + '/' + name, PATH + '/load.xlsx')
    wb = openpyxl.load_workbook(PATH + '/load.xlsx')
    ws = wb.active
    inserted_data[0] = ws['A' + str(ws.max_row)].value + 1
    inserted_data[1] = str(datetime.now().year) + '.' + str(str(datetime.now().strftime("%m"))) + '.' + str(datetime.now().strftime("%d"))
    if inserted_data[0] > 5000:
        wb.close()
        for i in range(2, 9):
            if str(datetime.now().year) + '_' + str(cemetery) + '_' + str(i) + '.xlsx' in tabs_list:
                wb = openpyxl.load_workbook(str(datetime.now().year) + '_' + str(cemetery) + '_' + str(i) + '.xlsx')
                ws = wb.active
                if ws['A' + str(ws.max_row)].value + 1 > 5000:
                    continue
                else:
                    add_row(parent_func, cemetery, inserted_data, str(datetime.now().year) + '_' + str(cemetery) + '_' + str(i) + '.xlsx')
                    break
            elif str(datetime.now().year) + '_' + str(cemetery) + '_' + str(i) + '.xlsx' not in tabs_list:
                create_new_excel(parent_func, str(datetime.now().year) + '_' + str(cemetery) + '_' + str(i) + '.xlsx', inserted_data)
                break                
        return 0
    ws.append(inserted_data)
    set_border(ws, 'A' + str(ws.max_row) + ':' + 'Q' + str(ws.max_row))
    
    df.loc[len(df)+1] = inserted_data
    change_label_bg_color(parent_func, inserted_data)
    table.delete(*table.get_children())
    table.insert('', tk.END, values=list(df.iloc[len(df)-1]))
    # Сохранение изменений
    wb.save(PATH + '/load.xlsx')
    shutil.copy(PATH + '/load.xlsx', PATH + '/' + name)

def add_fio_old():
    global table
    global age_label_old
    global last_name_label_add_old
    global name_label_add_old
    global number_label_old
    global entries_old
    global df
    global cemetery_label_old

    table.delete(*table.get_children())
    inserted_data = ['', '']
    for i in range(0, len(entries_old)-10):
        inserted_data.append(entries_old[i].get().strip().capitalize())
    for i in range(len(entries_old)-10, len(entries_old)):
        if i == 8:
            inserted_data.append(entries_old[i].get().strip().capitalize())
        else:
            inserted_data.append(entries_old[i].get().strip())
    try:
        inserted_data[5] = int(inserted_data[5])
        age = inserted_data[5]
    except:
        age = -1
        age_label_old['bg'] = '#dd6666'
        number_label_old['text'] = "Запись не была произведена!"
    if inserted_data[2] == '':
        last_name_label_add_old['bg'] = '#dd6666'
        number_label_old['text'] = "Запись не была произведена!"
    else:
        last_name_label_add_old['bg'] = '#ffffff'
    if inserted_data[3] == '':
        name_label_add_old['bg'] = '#dd6666'
        number_label_old['text'] = "Запись не была произведена!"
    else:
        name_label_add_old['bg'] = '#ffffff'
    if age < 0:
        age_label_old['bg'] = '#dd6666'
        number_label_old['text'] = "Запись не была произведена!"
    else:
        age_label_old['bg'] = '#ffffff'
    if 'Покровское' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    elif 'Южное' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    elif 'Старопокровское' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    else:
        cemetery_is_set = False
        cemetery_label_old['bg'] = '#dd6666'
        number_label_old['text'] = "Запись не была произведена!"
    
    if inserted_data[2] != '' and inserted_data[3] != '' and age > -1 and cemetery_is_set:
        add_row('old', cemetery, inserted_data, '2011' + '_' + str(cemetery) + '.xlsx')

def add_fio():
    global table
    global age_label
    global last_name_label_add
    global name_label_add
    global number_label
    global entries
    global entries_old
    global df
    global cemetery_label

    table.delete(*table.get_children())
    inserted_data = ['', '']
    for i in range(0, len(entries)-13):
        inserted_data.append(entries[i].get().strip().capitalize())
    for i in range(len(entries)-13, len(entries)-3):
        if i == 8:
            inserted_data.append(entries[i].get().strip().capitalize())
        else:
            inserted_data.append(entries[i].get().strip())

    try:
        inserted_data[5] = int(inserted_data[5])
        age = inserted_data[5]
    except:
        age = -1
        age_label['bg'] = '#dd6666'
        number_label['text'] = "Запись не была произведена!"
    if inserted_data[2] == '':
        last_name_label_add['bg'] = '#dd6666'
        number_label['text'] = "Запись не была произведена!"
    else:
        last_name_label_add['bg'] = '#ffffff'
    if inserted_data[3] == '':
        name_label_add['bg'] = '#dd6666'
        number_label['text'] = "Запись не была произведена!"
    else:
        name_label_add['bg'] = '#ffffff'
    if age < 0:
        age_label['bg'] = '#dd6666'
        number_label['text'] = "Запись не была произведена!"
    else:
        age_label['bg'] = '#ffffff'
    if 'Покровское' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    elif 'Южное' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    elif 'Старопокровское' == inserted_data[10]:
        cemetery = inserted_data[10]
        cemetery_is_set = True
    else:
        cemetery_is_set = False
        cemetery_label['bg'] = '#dd6666'
        number_label['text'] = "Запись не была произведена!"
    
    if inserted_data[2] != '' and inserted_data[3] != '' and age > -1 and cemetery_is_set:
        add_row('now', cemetery, inserted_data, str(datetime.now().year) + '_' + str(cemetery) + '.xlsx')       

def check_containing(need_name, name):
    if (len(need_name) > 0 and  need_name in name or len(need_name) > 0 and  name in need_name) or len(need_name) == 0:
        return 1
    else:
        return 0

def find_fio():
    global table
    global entries
    global df
    table.delete(*table.get_children())
    names = list(df['Имя умершего'])
    last_names = list(df['Фамилия умершего'])
    surnames = list(df['Отчество умершего'])
    need_name = entries[len(entries)-3].get().strip().capitalize()
    need_last_name = entries[len(entries)-2].get().strip().capitalize()
    need_surname = entries[len(entries)-1].get().strip().capitalize()
    for i in range(1, len(names)):
        if check_containing(need_name, str(names[i])) + check_containing(need_last_name, str(last_names[i])) + check_containing(need_surname, str(surnames[i])) == 3:
            table.insert('', tk.END, values=list(df.iloc[i]))    


def open_main_window():
    # Создание главного окна
    global name_entry
    global last_name_entry
    global surname_entry
    global table
    global entries
    global entries_old
    global age_label
    global last_name_label_add
    global name_label_add
    global number_label
    global number_label_old
    global frame_add
    global frame_find
    global frame_add_old
    global root
    global cemetery_label
    global age_label_old
    global last_name_label_add_old
    global name_label_add_old
    global cemetery_label_old
    root = tk.Tk()
    root.title("Работа с эксель таблицами")
    root.geometry("1250x440")
    label_bg_colors = 0xffffff
    PATH_output_file = ''

    notebook = ttk.Notebook()
    notebook.grid(row=1, column=0, sticky='w')
    
    # создаем пару фреймвов
    frame_find = ttk.Frame(notebook)
    frame_add = ttk.Frame(notebook)
    frame_add_old = ttk.Frame(notebook)

    frame_find.grid()
    frame_add.grid()
    frame_add_old.grid()
    # добавляем фреймы в качестве вкладок
    notebook.add(frame_find, text="Поиск по ФИО")
    notebook.add(frame_add, text="Добавление текущих лет")
    notebook.add(frame_add_old, text="Добавление старых лет")

    last_name_label = tk.Label(frame_find,
                    text="Фамилия",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    last_name_label.grid(row=3, column=0, sticky='we')
    last_name_entry = tk.Entry(frame_find, font= 20)
    last_name_entry.grid(row=3, column=1,sticky='we')

    name_label = tk.Label(frame_find,
                    text="Имя",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    name_label.grid(row=4, column=0, sticky='we')
    name_entry = tk.Entry(frame_find, font= 20)
    name_entry.grid(row=4, column=1,sticky='we')

    surname_label = tk.Label(frame_find,
                    text="Отчество",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=5, column=0, sticky='we')
    surname_entry = tk.Entry(frame_find, font= 20)
    surname_entry.grid(row=5, column=1,sticky='we')

    last_name_label_add = tk.Label(frame_add,
                    text="Фамилия",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    last_name_label_add.grid(row=2, column=0, sticky='we')
    last_name_entry_add = tk.Entry(frame_add, font= 20)
    last_name_entry_add.grid(row=2, column=1,sticky='we')

    name_label_add = tk.Label(frame_add,
                    text="Имя",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    name_label_add.grid(row=3, column=0, sticky='we')
    name_entry_add = tk.Entry(frame_add, font= 20)
    name_entry_add.grid(row=3, column=1,sticky='we')

    surname_label_add = tk.Label(frame_add,
                    text="Отчество",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=0, sticky='we')
    surname_entry_add = tk.Entry(frame_add, font= 20)
    surname_entry_add.grid(row=4, column=1,sticky='we')

    age_label = tk.Label(frame_add,
                    text="Возраст умершего",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    age_label.grid(row=5, column=0, sticky='we')
    age_entry = tk.Entry(frame_add, font= 20)
    age_entry.grid(row=5, column=1,sticky='we')

    death_date_label = tk.Label(frame_add,
                    text="Дата смерти",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=0, sticky='we')
    death_date_entry = tk.Entry(frame_add, font= 20)
    death_date_entry.grid(row=6, column=1,sticky='we')

    certificate_number_label = tk.Label(frame_add,
                    text="Номер свидетельства",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=2, column=2, sticky='we')
    certificate_number_entry = tk.Entry(frame_add, font= 20)
    certificate_number_entry.grid(row=2, column=3,sticky='we')

    given_label = tk.Label(frame_add,
                    text="Каким ЗАГСом выдано",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=3, column=2, sticky='we')
    given_entry = tk.Entry(frame_add, font= 20)
    given_entry.grid(row=3, column=3,sticky='we')

    burial_label = tk.Label(frame_add,
                    text="Дата погребения",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=2, sticky='we')
    burial_entry = tk.Entry(frame_add, font= 20)
    burial_entry.grid(row=4, column=3,sticky='we')

    cemetery_label = tk.Label(frame_add,
                    text="Кладбище",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    cemetery_label.grid(row=5, column=2, sticky='we')
    cemetery_entry = tk.Entry(frame_add, font= 20)
    cemetery_entry.grid(row=5, column=3,sticky='we')

    sector_number_label = tk.Label(frame_add,
                    text="Номер сектора могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=2, sticky='we')
    sector_number_entry = tk.Entry(frame_add, font= 20)
    sector_number_entry.grid(row=6, column=3,sticky='we')

    row_label = tk.Label(frame_add,
                    text="Номер ряда могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=2, column=4, sticky='we')
    row_entry = tk.Entry(frame_add, font= 20)
    row_entry.grid(row=2, column=5,sticky='we')

    place_number_label = tk.Label(frame_add,
                    text="Номер места могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=3, column=4, sticky='we')
    place_number_entry = tk.Entry(frame_add, font= 20)
    place_number_entry.grid(row=3, column=5,sticky='we')

    size_label = tk.Label(frame_add,
                    text="Размер отведенного участка",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=4, sticky='we')
    size_entry = tk.Entry(frame_add, font= 20)
    size_entry.grid(row=4, column=5,sticky='we')

    digger_label = tk.Label(frame_add,
                    text="Фамилия землекопа",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=5, column=4, sticky='we')
    digger_entry = tk.Entry(frame_add, font= 20)
    digger_entry.grid(row=5, column=5,sticky='we')

    responsibility_label = tk.Label(frame_add,
                    text="Взявший ответственность",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=4, sticky='we')
    responsibility_entry = tk.Entry(frame_add, font= 20)
    responsibility_entry.grid(row=6, column=5,sticky='we')

    button_to_find = tk.Button(frame_find, 
                            text = "Найти",
                            command = find_fio,
                            ).grid(row=7,sticky='we', columnspan=2)

    button_to_add = tk.Button(frame_add, 
                            text = "Добавить",
                            command = add_fio,
                            ).grid(row=7, column=0,sticky='we', columnspan=7)
    



    last_name_label_add_old = tk.Label(frame_add_old,
                    text="Фамилия",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    last_name_label_add_old.grid(row=2, column=0, sticky='we')
    last_name_entry_add_old = tk.Entry(frame_add_old, font= 20)
    last_name_entry_add_old.grid(row=2, column=1,sticky='we')

    name_label_add_old = tk.Label(frame_add_old,
                    text="Имя",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    name_label_add_old.grid(row=3, column=0, sticky='we')
    name_entry_add_old = tk.Entry(frame_add_old, font= 20)
    name_entry_add_old.grid(row=3, column=1,sticky='we')

    surname_label_add_old = tk.Label(frame_add_old,
                    text="Отчество",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=0, sticky='we')
    surname_entry_add_old = tk.Entry(frame_add_old, font= 20)
    surname_entry_add_old.grid(row=4, column=1,sticky='we')

    age_label_old = tk.Label(frame_add_old,
                    text="Возраст умершего",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    age_label_old.grid(row=5, column=0, sticky='we')
    age_entry_old = tk.Entry(frame_add_old, font= 20)
    age_entry_old.grid(row=5, column=1,sticky='we')

    death_date_label_old = tk.Label(frame_add_old,
                    text="Дата смерти",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=0, sticky='we')
    death_date_entry_old = tk.Entry(frame_add_old, font= 20)
    death_date_entry_old.grid(row=6, column=1,sticky='we')

    certificate_number_label_old = tk.Label(frame_add_old,
                    text="Номер свидетельства",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=2, column=2, sticky='we')
    certificate_number_entry_old = tk.Entry(frame_add_old, font= 20)
    certificate_number_entry_old.grid(row=2, column=3,sticky='we')

    given_label_old = tk.Label(frame_add_old,
                    text="Каким ЗАГСом выдано",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=3, column=2, sticky='we')
    given_entry_old = tk.Entry(frame_add_old, font= 20)
    given_entry_old.grid(row=3, column=3,sticky='we')

    burial_label_old = tk.Label(frame_add_old,
                    text="Дата погребения",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=2, sticky='we')
    burial_entry_old = tk.Entry(frame_add_old, font= 20)
    burial_entry_old.grid(row=4, column=3,sticky='we')

    cemetery_label_old = tk.Label(frame_add_old,
                    text="Кладбище",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    )
    cemetery_label_old.grid(row=5, column=2, sticky='we')
    cemetery_entry_old = tk.Entry(frame_add_old, font= 20)
    cemetery_entry_old.grid(row=5, column=3,sticky='we')

    sector_number_label_old = tk.Label(frame_add_old,
                    text="Номер сектора могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=2, sticky='we')
    sector_number_entry_old = tk.Entry(frame_add_old, font= 20)
    sector_number_entry_old.grid(row=6, column=3,sticky='we')

    row_label_old = tk.Label(frame_add_old,
                    text="Номер ряда могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=2, column=4, sticky='we')
    row_entry_old = tk.Entry(frame_add_old, font= 20)
    row_entry_old.grid(row=2, column=5,sticky='we')

    place_number_label_old = tk.Label(frame_add_old,
                    text="Номер места могилы",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=3, column=4, sticky='we')
    place_number_entry_old = tk.Entry(frame_add_old, font= 20)
    place_number_entry_old.grid(row=3, column=5,sticky='we')

    size_label_old = tk.Label(frame_add_old,
                    text="Размер отведенного участка",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=4, column=4, sticky='we')
    size_entry_old = tk.Entry(frame_add_old, font= 20)
    size_entry_old.grid(row=4, column=5,sticky='we')

    digger_label_old = tk.Label(frame_add_old,
                    text="Фамилия землекопа",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=5, column=4, sticky='we')
    digger_entry_old = tk.Entry(frame_add_old, font= 20)
    digger_entry_old.grid(row=5, column=5,sticky='we')

    responsibility_label_old = tk.Label(frame_add_old,
                    text="Взявший ответственность",
                    bg= '#' + str(hex(label_bg_colors))[2:],
                    font= 20
                    ).grid(row=6, column=4, sticky='we')
    responsibility_entry_old = tk.Entry(frame_add_old, font= 20)
    responsibility_entry_old.grid(row=6, column=5,sticky='we')

    button_to_add_old = tk.Button(frame_add_old, 
                            text = "Добавить",
                            command = add_fio_old,
                            ).grid(row=7, column=0,sticky='we', columnspan=7)

    entries = [last_name_entry_add, name_entry_add, surname_entry_add, age_entry, death_date_entry, certificate_number_entry, given_entry, 
            burial_entry, cemetery_entry, sector_number_entry, row_entry, place_number_entry, size_entry, digger_entry, responsibility_entry,  name_entry,  last_name_entry,  surname_entry]

    entries_old = [last_name_entry_add_old, name_entry_add_old, surname_entry_add_old, age_entry_old, death_date_entry_old, certificate_number_entry_old, given_entry_old, 
            burial_entry_old, cemetery_entry_old, sector_number_entry_old, row_entry_old, place_number_entry_old, size_entry_old, digger_entry_old, responsibility_entry_old]
    
    table = ttk.Treeview(columns=headers_list, show="headings")


    xscrollbar = ttk.Scrollbar(root,orient=HORIZONTAL,command=table.xview)
    xscrollbar.grid(row=10, column=0, sticky=EW, columnspan=6)
    table.configure(xscrollcommand=xscrollbar.set)

    yscrollbar = ttk.Scrollbar(root,orient=VERTICAL,command=table.yview)
    yscrollbar.grid(row=9, column=8, sticky=NS)
    table.configure(yscrollcommand=yscrollbar.set)

    frame_find.grid_columnconfigure(0, minsize=175)
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(0, weight=1)
        
    table.grid(row=9, column=0, sticky="nsew", columnspan=6)

    root.mainloop()

# Создание и запуск потока для выполнения вычислений
calc_thread = threading.Thread(target=set_df)
calc_thread.start()

# Открытие окна приложения
open_main_window()